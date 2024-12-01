from tqdm import tqdm
import os
import xlwings as xw
from openpyxl import Workbook
from openpyxl.styles import Font
import time
from openpyxl.utils import get_column_letter

def extract_data(file_path):
    try:
        data = {}
        with xw.App(visible=False) as app:
            wb = app.books.open(file_path)
            if 'Signaletiq' in [sheet.name for sheet in wb.sheets]:
                sheet_signaletiq = wb.sheets['Signaletiq']
                bank_name = sheet_signaletiq.range('B4').value
                data['bank_name'] = bank_name if bank_name else 'Unknown Bank'
            else:
                print(f"'Signaletiq' sheet not found in {file_path}")
                wb.close()
                return None
            cell_ranges = {
                'actif': [4, 12, 13, 14, 15, 16, 17, 18, 19, 21, 28, 31, 42, 35, 50],
                'passif': [3, 18, 19, 20, 21, 22, 23, 30, 34, 40],
                'cpte de result Charges': [3, 11, 14, 28, 29],
                'cpte de result Produits': [3, 12, 20, 24],
            }
            for sheet_name, rows in cell_ranges.items():
                if sheet_name in [sheet.name for sheet in wb.sheets]:
                    sheet = wb.sheets[sheet_name]
                    data[sheet_name] = []
                    for row in rows:
                        key = sheet.range(f'B{row}').value
                        value = sheet.range(f'C{row}').value
                        if key is not None:
                            data[sheet_name].append((key, value))
                else:
                    print(f"'{sheet_name}' sheet not found in {file_path}")
            wb.close()
        return data
    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        return None
def get_next_filename(output_folder, base_filename):
    file_num = 1
    while os.path.exists(os.path.join(output_folder, f"{base_filename}{file_num}.xlsx")):
        file_num += 1
    return f"{base_filename}{file_num}.xlsx"
def consolidate_data(files, output_folder, base_filename):
    consolidated_data = {}
    for file in files:
        data = extract_data(file)
        if not data:
            continue
        bank_name = data.pop('bank_name')
        if bank_name not in consolidated_data:
            consolidated_data[bank_name] = {sheet: [] for sheet in data.keys()}
        for section, items in data.items():
            for key, value in items:
                if key not in [item[0] for item in consolidated_data[bank_name][section]]:
                    consolidated_data[bank_name][section].append((key, value))
    if not consolidated_data:
        print("No data extracted. Check input files and cell references.")
        return
    wb = Workbook()
    for section in ['actif', 'passif', 'cpte de result Charges', 'cpte de result Produits']:
        ws = wb.create_sheet(title=section)
        banks = list(consolidated_data.keys())
        headers = ["Item"] + banks + ["Total"]
        ws.append(headers)
        all_items = []
        for bank_data in consolidated_data.values():
            all_items.extend([item[0] for item in bank_data.get(section, [])])
        all_items = list(dict.fromkeys(all_items))  
        row_index = 2 
        for item in all_items:
            if section == 'actif' and row_index in range(4, 11): 
                row_item = f"- {item}" if item else item
            elif section == 'passif' and row_index in range(4, 8):
                row_item = f"- {item}" if item else item
            else:
                row_item = item
            row = [row_item]
            total = 0
            for bank in banks:
                value = next((v for k, v in consolidated_data[bank][section] if k == item), 0)
                row.append(value)
                total += value if isinstance(value, (int, float)) else 0
            row.append(total)
            ws.append(row)
            if row_item.startswith("-"):
                ws.row_dimensions[row_index].hidden = True 
                # ws.row_dimensions[row_index - 1].outline_level = 1
                ws.row_dimensions[row_index].outline_level = 1  
                ws.row_dimensions[row_index].group = True  
            row_index += 1
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    output_file = get_next_filename(output_folder, base_filename) 
    wb.save(os.path.join(output_folder, output_file))
    print(f"Consolidated data saved to {os.path.join(output_folder, output_file)}")
    return output_file    

def main():
    input_folder = "test"  
    output_folder = "et"  
    base_filename = "et"  
    files = [
        os.path.join(input_folder, f) 
        for f in os.listdir(input_folder) 
        if f.endswith('.xlsx') and not f.startswith('~$')
    ]
    if not files:
        print("No valid Excel files found in the specified folder.")
        return
    print(f"Found {len(files)} files to process.")
    with tqdm(total=100, desc="Consolidating Data", unit="%", ncols=100) as pbar:
        for i in range(1, 101):  
            time.sleep(0.05)  
            pbar.update(1)  
    consolidate_data(files, output_folder, base_filename)
    print("Data consolidation complete!")

if __name__ == "__main__":
    main()
